from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook,drawing
from openpyxl.utils import get_column_letter
import json
import pandas as pd
from datetime import datetime
import datetime as dt
import matplotlib.pyplot as plt

class Initiate_Sheets:

    def __init__(self, session_filename, start_time, end_time):
        self.max_value = 0
        self.min_value = 0
        self.record_amount = 0
        self.data = None
        self.daily_amount = {}

        self.wb = Workbook()
        self.overall_ws = self.wb.active
        self.overall_ws.title='Overall'

        self.statistic_ws = self.wb.create_sheet()
        self.statistic_ws.title='FAQ_Statistic'

        self.agent_ws = self.wb.create_sheet()
        self.agent_ws.title='Agent_Performance'

        self.utilization_ws = self.wb.create_sheet()
        self.utilization_ws.title='Utilization'

        self.thin = Side(border_style="thin", color="000000")  # border in black
        self.reformatted_session_data_dict = {}
        self.datetime_list_5min ={}
        self.Total_IVR_Session = 0 
        self.Total_Agent_Session = 0
        # self.earliest_date = None
        # self.latest_date = None
        self.earliest_date = self.date_time_string2object(start_time)  #to be set from API 
        self.latest_date   = self.date_time_string2object(end_time)  #to be set from API
        self.datelist = None
        self.agent_performance_dict= {}
        self.overall_agent_performance_dict ={}
        #self.agents = ['BEN', 'ROGER']
        self.agents = set()
        self.draw_plot_at_row = 25

        self.obtain_session_data(session_filename)
        self.get_agent_names()
        self.compute_overall_n_utilization()
        self.compute_agent_performance()
        self.compute_overall_agents_performance()

        self.colors = ['red', 'blue', 'green', 'yellow', 'orange', 'purple', 'pink', 'brown', 'olive', 'cyan', 'skyblue', 'lightgreen', 'salmon', 'tomato', 'darkviolet', 'gold', 'lime', 'darkgreen', 'peru', 'darkcyan', 'violet', 'navy']  

        

    def obtain_session_data(self, session_filename):
        #reformatted_session_data_dict = {}
        with open(session_filename, encoding='utf-8') as file_object:
            session_data = json.load(file_object)

        # self.earliest_date = self.date_time_string2object(session_data[0]['datetime'])
        # self.latest_date   = self.date_time_string2object(session_data[0]['datetime'])

        for session in session_data:
            # # ============== get ealiest_date, latest_date ==============
            # date_time_obj = self.date_time_string2object(session['datetime'])
            # if date_time_obj < self.earliest_date:
            #     self.earliest_date = date_time_obj
            # if date_time_obj > self.latest_date:
            #     self.latest_date = date_time_obj

            #========== reformatted the session data ==========
            if not session['session_id'] in self.reformatted_session_data_dict:
                self.reformatted_session_data_dict[session['session_id']] = []
            mysession = {}
            mysession['datetime'] = session['datetime']
            mysession['type'] = session['type']
            mysession['msg'] = session['msg']
            self.reformatted_session_data_dict[session['session_id']].append(mysession)
            
        # sample of reformatted_session_data_dict
        # reformatted_session_data_dict = {'UGU273': [{'datetime': '2021-01-09 14:00:10', 'type': 'FAQ', 'msg': 'FAQ_ID_ru783jdsl2943'}, {'datetime': '2021-01-09 14:01:00', 'type': 'COMMAND', 'msg': 'IVR_SESSION_END'}], 'YRU2566': [{'datetime': '2021-02-10 18:00:02', 'type': 'FAQ', 'msg': 'FAQ_ID_3969030'}, {'datetime': '2021-02-10 18:00:05', 'type': 'FAQ', 'msg': 'FAQ_ID_3969031'}, {'datetime': '2021-02-10 18:02:00', 'type': 'COMMAND', 'msg': 'TRANSFER_TO_EXTENSION'}, {'datetime': '2021-02-10 18:02:02', 'type': 'COMMAND', 'msg': 'PICK_UP_BY_BEN'}, {'datetime': '2021-02-10 18:05:05', 'type': 'COMMAND', 'msg': 'CALL_END'}], 'ERT899': [{'datetime': '2021-02-10 19:00:02', 'type': 'FAQ', 'msg': 'FAQ_ID_3969030'}, {'datetime': '2021-02-10 19:00:05', 'type': 'FAQ', 'msg': 'FAQ_ID_3969031'}, {'datetime': '2021-02-10 19:02:00', 'type': 'COMMAND', 'msg': 'TRANSFER_TO_EXTENSION'}, {'datetime': '2021-02-10 19:04:00', 'type': 'COMMAND', 'msg': 'CALL_END'}]}

        #print(f"self.reformatted_session_data_dict: {self.reformatted_session_data_dict}")
        #print(f'earliest_date: {earliest_date}')
        #print(f'latest_date: {latest_date}')

        earliest_hour = self.earliest_date.hour
        latest_hour = self.latest_date.hour
        earliest_minute = self.earliest_date.minute
        latest_minute = self.latest_date.minute
        
        # print(f'earliest_hour: {earliest_hour}, earliest_minute: {earliest_minute}')
        # print(f'latest_hour: {latest_hour}, latest_minute: {latest_minute}')  
        
        # print(f'self.earliest_date: {self.earliest_date}')
        # print(f'self.latest_date: {self.latest_date}')
        self.datelist = self.get_date_in_range(self.date_time_object2string(self.earliest_date), self.date_time_object2string(self.latest_date))

        # print('get_date_in_range')
        # for j in self.datelist:
        #     date = j.strftime("%Y-%m-%d")
        #     print(date, type(date))

        my_24hr_list = self.get_24hr_in_every_5min()
        start_index = earliest_hour*12 + int(earliest_minute/5)  # since the interval is 5 mins
        end_index = len(self.datelist)*len(my_24hr_list) - (((23-latest_hour)*12)+ 12-int(latest_minute/5))

        datetime_list = []
        for m in self.datelist:
            date = m.strftime("%Y-%m-%d")
            for k in my_24hr_list:
                datetime_list.append(f'{date} {k}')

        # self.datetime_list_5min ={}
        for date_time in datetime_list[start_index:end_index+1]:
            self.datetime_list_5min[date_time] = {'ivr_calls':0, 'ivr_throu':0, 'agent_calls':0,'agent_throu':0}

        # print("------")
        # print("start line 112")
        # for i in self.datetime_list_5min:
        #     print(i)
        # print("line 112")


    def get_agent_names(self):
        for key, session in self.reformatted_session_data_dict.items():
            for item in session:
                if item['type'] == "COMMAND" and "PICK_UP_BY" in item['msg']:
                    name = item['msg'].replace("PICK_UP_BY_", "")
                    self.agents.add(name)

        print(f'ALL agents: {self.agents}')
    

    def classify_timeslot_in_5min(self, date_time_str):
        # e.g. date_time_str is "2021-01-09 14:53:02"
        date_time_obj = datetime.strptime(date_time_str, "%Y-%m-%d %H:%M:%S")
        belong_minute = 5*(int(date_time_obj.minute/5))
        if belong_minute < 10:
            belong_minute = '0'+str(belong_minute)

        belong_to = f'{date_time_str.split(":")[0]}:{belong_minute}:00'
        return belong_to


    def compute_overall_n_utilization(self):
        self.Total_IVR_Session = len(self.reformatted_session_data_dict) 

        for _, session in self.reformatted_session_data_dict.items():  #session_id
            for item in session:
                #print(item)
                if item['type'] == "COMMAND" and 'IVR_SESSION_START' in item['msg']:
                    ivr_belonging_timeslot = self.classify_timeslot_in_5min(item['datetime'])
                    #print(f'ivr_belonging_timeslot: {ivr_belonging_timeslot}')
                    # assign to the dict
                    self.datetime_list_5min[ivr_belonging_timeslot]['ivr_calls'] += 1
                if item['type'] == "COMMAND" and 'PICK_UP_BY' in item['msg']:
                    self.Total_Agent_Session +=1
                    agent_belonging_timeslot = self.classify_timeslot_in_5min(item['datetime'])
                    self.datetime_list_5min[agent_belonging_timeslot]['agent_calls'] += 1
    
    
    def get_duration(self, time1, time2):
        # param: time1, time2 in str
        # return in sec (int)
        d1 = datetime.strptime(time1, '%Y-%m-%d %H:%M:%S')
        d2 = datetime.strptime(time2, '%Y-%m-%d %H:%M:%S')

        # # return in timedelta (str)
        # diff = d2-d1
        # print(diff, type(diff)) # <class 'datetime.timedelta'>
        # return f'{diff}'

        # if in sec is needed, use the following
        diff_sec = (d2-d1).total_seconds()
        # print(diff_sec, type(diff_sec))  # <class 'float'>
        return int(diff_sec)


    def calculate_agent_4_timeslot(self, agent_name, date_time):
        mydict = {}
        # return example:
        # #mydict['Agent_'+str(order+1)] = agent_name
        # mydict['Agent'] = agent_name
        # mydict['Number_of_Calls_Handled'] = 10
        # mydict['Average_Call_Duration'] = 20
        # mydict['Average_Wait_Time'] = 30
        # mydict['First_Call_Resolution_Rate'] = 100
        # mydict['Available_Time'] = 100

        average_call_duration = 0
        average_wait_time = 0
        total = 0

        target_date_hour = date_time.split(":")[0]
        # iterate the reformatted dict, 
        #   if the datetime belongs the one I am looking for, 
        #       if PICK_UP_BY_+agent_name.upper() exist in the session:
        #          average_call_duration = (average_call_duration*total+[cal_duration])/(total+1)
        #          average_wait_time = (average_wait_time*total+[cal_wait])/(total+1)
        #          total +=1                
        #       else:
        #           skip
        #   else:
        #       skip the session
        for _, item_ls in self.reformatted_session_data_dict.items():
            #new_ls = sorted(item_ls, key=lambda k:k['datetime'])

            # sort the list by date_time in timestamp
            new_ls = sorted(item_ls, key=lambda k:datetime.timestamp(datetime.strptime(k['datetime'], "%Y-%m-%d %H:%M:%S")))
            
            # should look for the record with TRANSFER_TO_EXTENSION
            valid_record = False
            for item in new_ls:
                if item["type"] == "COMMAND" and item["msg"] == "TRANSFER_TO_EXTENSION":
                    wait_start = item["datetime"]
                if item["type"] == "COMMAND" and item["msg"] == "PICK_UP_BY_"+agent_name.upper() and target_date_hour in item["datetime"]:
                    valid_record = True
                    wait_end = item["datetime"]
            
            if valid_record and "CALL_END" in new_ls[-1]["msg"]:
                # why check valid_record? As the IVR session could be ended without agent picking up the call and in such case it wont be counted as agent data
                call_end = new_ls[-1]["datetime"]

                average_call_duration = (average_call_duration*total+self.get_duration(wait_end, call_end))/(total+1)
                average_wait_time = (average_wait_time*total+self.get_duration(wait_start,wait_end))/(total+1)
                total +=1

        mydict['Agent'] = agent_name.upper()
        mydict['Number_of_Calls_Handled'] = total
        mydict['Average_Call_Duration'] = average_call_duration
        mydict['Average_Wait_Time'] = average_wait_time
        mydict['First_Call_Resolution_Rate'] = 'N/A'
        mydict['Available_Time'] = 'N/A'
        return mydict


    def compute_agent_performance(self):
        earliest_hour = self.earliest_date.hour
        latest_hour = self.latest_date.hour

        my_24 = []  # '00:00 - 00:59'
        hour = 0
        while hour <=23:
            hour_ = str(hour)
            if hour < 10:
                hour_ = '0'+str(hour)
            my_24.append(f'{hour_}:00-{hour_}:59')
            hour +=1        

        final_datetime_list = []
        for m in self.datelist:
            date = m.strftime("%Y-%m-%d")
            for hr in my_24:
                final_datetime_list.append(date+' '+hr)
        #print(final_datetime_list)
        #final_datetime_list = final_datetime_list[earliest_hour:len(final_datetime_list)-(23-latest_hour)]
        #print(final_datetime_list[earliest_hour:len(final_datetime_list)-(23-latest_hour)])
        
        for date_time in final_datetime_list[earliest_hour:len(final_datetime_list)-(23-latest_hour)]:
            #print(f'date_time: {date_time}')
            self.overall_agent_performance_dict[date_time] = {} # initiate the dict for overall performance of all agents

            self.agent_performance_dict[date_time] = []
            for _, agent_name in enumerate(self.agents):
                #agent_detail = self.calculate_agent_4_timeslot(i, agent_name, date_time)
                agent_detail = self.calculate_agent_4_timeslot(agent_name, date_time)
                self.agent_performance_dict[date_time].append(agent_detail)

        #print(self.agent_performance_dict)


    def overall_extension_agent_4_timeslot(self, date_time):
        # Overall Total Number of Call = all session that involve TRANSFER_TO_EXTENSION
        # Overall Number of Call Handled = all session that involve TRANSFER_TO_EXTENSION and PICK_UP_BY_name
        amount_extension, amount_agent = 0, 0
        target_date_hour = date_time.split(":")[0]
        for _, session in self.reformatted_session_data_dict.items():
            for action in session:
                if action["type"] == 'COMMAND' and action["msg"] == 'TRANSFER_TO_EXTENSION' and action["datetime"].split(":")[0] == target_date_hour:
                    amount_extension +=1
                elif action["type"] == 'COMMAND' and 'PICK_UP_BY' in action["msg"] and action["datetime"].split(":")[0] == target_date_hour:
                    amount_agent += 1
        return amount_extension, amount_agent


    def compute_overall_agents_performance(self):
        #print(self.agent_performance_dict.items())
        for date_time, item in self.agent_performance_dict.items():
            amount_extension, amount_agent = self.overall_extension_agent_4_timeslot(date_time)
            average_call_duration = 0
            average_wait_time = 0
            sum_calls = 0
            for agent in item:
                # Average_Call_Duration format only numeric
                if sum_calls+agent["Number_of_Calls_Handled"] != 0:   # to prevent DivideZeroError
                    average_call_duration = (average_call_duration*sum_calls + agent["Average_Call_Duration"]*agent["Number_of_Calls_Handled"])/(sum_calls+agent["Number_of_Calls_Handled"])
                    average_wait_time     = (average_wait_time*sum_calls + agent["Average_Wait_Time"]*agent["Number_of_Calls_Handled"])/(sum_calls+agent["Number_of_Calls_Handled"])
                    sum_calls = sum_calls + agent["Number_of_Calls_Handled"]

            self.overall_agent_performance_dict[date_time]['Total_Number_of_Call'] = amount_extension
            self.overall_agent_performance_dict[date_time]['Number_of_Calls_Handled'] = amount_agent  # amount_agent should = sum_calls
            self.overall_agent_performance_dict[date_time]['Average_Call_Duration'] = average_call_duration
            self.overall_agent_performance_dict[date_time]['Average_Wait_Time'] = average_wait_time


    def set_border_for_all(self, ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        side = Side(border_style='thin', color="FF000000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border
        

    def get_max_min(self, data):
        global max_value
        global min_value

        #max_value = 214
        min_value = 0
        max_value = 0
        for _, value in self.daily_amount.items():
            if value['total_triggered'] > max_value:
                max_value = value['total_triggered']
            if value['total_responsed'] > max_value:
                max_value = value['total_responsed']

        return max_value, min_value


    def get_daily_amount_triggered_responsed(self, data):
        ''' sample return
        daily_amount = {
            time1:{'total_triggered':100, 'total_responsed':50},
            time2:{'total_triggered':110, 'total_responsed':51},
            time3:{'total_triggered':102, 'total_responsed':52},
        }
        '''

        for i in data:
            for record in i['records']:
                if not record['date_time'] in self.daily_amount:
                    self.daily_amount[record['date_time']] = {}
                    self.daily_amount[record['date_time']]['total_triggered'] = 0
                    self.daily_amount[record['date_time']]['total_responsed'] = 0
                self.daily_amount[record['date_time']]['total_triggered'] += record['triggered']
                self.daily_amount[record['date_time']]['total_responsed'] += record['responsed']

        return self.daily_amount


    def get_hex_color(self, value, primary_color='red'):
        #max_value, min_value = get_max_min()
        color_level = round((max_value-value)/(max_value - min_value)*255)
        #print(f'color_level: {color_level}')

        if primary_color == 'red':
            return "FF{val}{val}".format(val=hex(color_level)[2:].zfill(2))
        elif primary_color == 'green':
            return "{val}FF{val}".format(val=hex(color_level)[2:].zfill(2))
        elif primary_color == 'blue':
            return "{val}{val}FF".format(val=hex(color_level)[2:].zfill(2))
        else:
            # by default - RED
            return "FF{val}{val}".format(val=hex(color_level)[2:].zfill(2))


    def create_statistic_ws(self, filename):

        with open(filename, encoding='utf-8') as file_object:
            data = json.load(file_object)
        
        #global record_amount
        self.record_amount = len(data)
        
        self.get_daily_amount_triggered_responsed(data)
        self.get_max_min(data)

        # ===========================================================================
        # CREATE COLUMN HEADERS FOR FAQ_id, FAQ_Path
        self.statistic_ws['A4'] = 'FAQ_id'
        self.statistic_ws['B4'] = 'FAQ_Path'
        self.statistic_ws.column_dimensions['A'].width = 8.43
        self.statistic_ws.column_dimensions['B'].width = 115.43

        # ===========================================================================
        # CREATE COLUMN HEADERS FOR DATA_TIME, Triggered, Responsed
        # starting from column 5 which Column E
        counter = 4
        for item in data[0]["records"]:
            
            column_letter_left = get_column_letter(counter+1)
            #print(f'column_letter_left:{column_letter_left}')
            counter +=2
            column_letter_right = get_column_letter(counter)

            # set width
            self.statistic_ws.column_dimensions[column_letter_left].width = 10
            self.statistic_ws.column_dimensions[column_letter_right].width = 10

            self.statistic_ws.merge_cells(column_letter_left+'2:'+column_letter_right+'2')  # starting from row 2

            self.statistic_ws[column_letter_left+'2'].value = item['date_time']

            self.set_border(self.statistic_ws, column_letter_left+'2:'+column_letter_right+'2')
            self.statistic_ws[column_letter_left+'2'].alignment = Alignment(horizontal="center", vertical="center")
        
            self.statistic_ws[column_letter_left+'3'].value  = 'Triggered'
            self.statistic_ws[column_letter_right+'3'].value = 'Responsed'
            #statistic_ws[column_letter_left+'3'].border = Border(left=thin, right=thin, bottom=thin)
            #statistic_ws[column_letter_right+'3'].border = Border(left=thin,right=thin, bottom=thin)

            self.statistic_ws[column_letter_left+'3'].alignment = Alignment(horizontal="center", vertical="center")
            self.statistic_ws[column_letter_right+'3'].alignment = Alignment(horizontal="center", vertical="center")

            #================== filling daily amount of total triggered and responsed ===============
            self.statistic_ws[column_letter_left+'4'].value  = self.daily_amount[item['date_time']]['total_triggered']  
            self.statistic_ws[column_letter_right+'4'].value = self.daily_amount[item['date_time']]['total_responsed']  
            self.statistic_ws[column_letter_left+'4'].fill  = PatternFill("solid", fgColor=self.get_hex_color(self.daily_amount[item['date_time']]['total_triggered']))
            self.statistic_ws[column_letter_right+'4'].fill = PatternFill("solid", fgColor=self.get_hex_color(self.daily_amount[item['date_time']]['total_responsed']))

            #============== styling - set borders ===================
            self.set_border_for_all(self.statistic_ws, column_letter_left+'3:'+column_letter_right+'4')
            if self.record_amount > 0:
                self.set_border(self.statistic_ws, column_letter_left+'5:'+column_letter_right+str(4+self.record_amount))

                self.set_border(self.statistic_ws,'A5:A'+str(4+self.record_amount))
                self.set_border(self.statistic_ws,'B5:B'+str(4+self.record_amount))
                self.set_border(self.statistic_ws,'C5:D'+str(4+self.record_amount))

            self.set_border(self.statistic_ws, 'A3:D3')
            self.set_border(self.statistic_ws, 'C4:D4')
            self.set_border_for_all(self.statistic_ws, 'A4:B4')

        # ===========================================================================
        # CREATE ROW content
        for row, item in enumerate(data):
            self.statistic_ws['A'+str(row+5)] = item['FAQ_id']
            self.statistic_ws['B'+str(row+5)] = item['FAQ_Path']
            column_counter = 4
            daily_sum_triggered = 0
            daily_sum_responsed = 0

            for record in item['records']:
                column_letter_left = get_column_letter(column_counter+1)
                #print(f'column_letter_left:{column_letter_left}')
                column_counter +=2
                column_letter_right = get_column_letter(column_counter)

                self.statistic_ws[column_letter_left+str(row+5)] = record['triggered']
                self.statistic_ws[column_letter_right+str(row+5)] = record['responsed']

                if type(record['triggered']) is int:
                    self.statistic_ws[column_letter_left+str(row+5)].fill = PatternFill("solid", fgColor=self.get_hex_color(record['triggered']))
                    daily_sum_triggered += record['triggered']
                else:
                    print(f"record['triggered']: {record['triggered']} is NOT int")
                if type(record['responsed']) is int:
                    self.statistic_ws[column_letter_right+str(row+5)].fill = PatternFill("solid", fgColor=self.get_hex_color(record['responsed']))
                    daily_sum_responsed += record['responsed']
                else:
                    print(f"record['responsed']: {record['responsed']} is NOT int")

            #=================== set daily_sum_triggered and daily_sum_responsed ====================
            self.statistic_ws['C'+str(row+5)] = daily_sum_triggered
            self.statistic_ws['D'+str(row+5)] = daily_sum_responsed
            self.statistic_ws['C'+str(row+5)].fill = PatternFill("solid", fgColor=self.get_hex_color(daily_sum_triggered))
            self.statistic_ws['D'+str(row+5)].fill = PatternFill("solid", fgColor=self.get_hex_color(daily_sum_responsed))


    def set_font(self, ws, cell, fontname='新細明體', size='11', bold=False):
        ws[cell].font = Font(name=fontname, size=size, bold=bold)


    def create_overall_ws(self):
        self.overall_ws.column_dimensions['A'].width = 8.43
        self.overall_ws.column_dimensions['B'].width = 40.14

        self.overall_ws['A1'].value = 'Management Information System Report'
        self.overall_ws['A2'].value = 'From'
        self.overall_ws['A3'].value = 'To'
        self.overall_ws['A4'].value = 'Channels'
        self.overall_ws['A5'].value = 'Agents'
        self.overall_ws['B2'].value = self.date_time_object2string(self.earliest_date) #'2019-11-01 00:00:00'
        self.overall_ws['B3'].value = self.date_time_object2string(self.latest_date)   #'2019-11-30 23:59:59'
        self.overall_ws['B4'].value = 'N/A'
        #self.overall_ws['B5'].value = '<<AGENT_1>>,<<AGENT_2>>,<<AGENT_3>>,<<AGENT_4>>,<<AGENT_5>>'
        self.overall_ws['B5'].value = ', '.join(self.agents)

        self.overall_ws.merge_cells('A1:B1')
        self.overall_ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        self.set_border(self.overall_ws, 'A2:A5')
        self.set_border(self.overall_ws, 'B2:B5')


        self.overall_ws['B8'].value = 'Overall Service Level Report'
        self.overall_ws.merge_cells('B8:C8')
        self.overall_ws['B8'].alignment = Alignment(horizontal="center", vertical="center")
        self.set_border(self.overall_ws, 'B8:C8')

        self.overall_ws['B9'].value = 'Total number of sessions received'
        self.overall_ws['B10'].value = 'Number of IVR Session'
        self.overall_ws['B11'].value = 'Number of Agent Session'
        self.set_border(self.overall_ws, 'B9:C11')


        self.overall_ws['B12'].value = 'IVR Session Performance'
        self.overall_ws.merge_cells('B12:C12')
        self.overall_ws['B12'].alignment = Alignment(horizontal="center", vertical="center")
        self.set_border(self.overall_ws, 'B12:C12')

        self.overall_ws['B13'].value = 'Number of Missed Calls'
        self.overall_ws['B14'].value = 'Number of Dropped Calls'
        self.overall_ws['B15'].value = 'Number of Lost Calls'
        self.overall_ws['B16'].value = 'Number of Short Calls'
        self.overall_ws['B17'].value = 'Average Wait Time (Sec)'
        self.overall_ws['B18'].value = 'Average Call Duration (Sec)'
        self.overall_ws['B19'].value = 'Call Abandon Rate (%)'
        self.overall_ws['B20'].value = 'First Call Resolution Rate'
        self.set_border(self.overall_ws, 'B13:C20')

        self.overall_ws['B21'].value = 'Call Center Performance'
        self.overall_ws.merge_cells('B21:C21')
        self.overall_ws['B21'].alignment = Alignment(horizontal="center", vertical="center")
        self.set_border(self.overall_ws, 'B21:C21')

        self.overall_ws['B22'].value = 'Number of Missed Calls'
        self.overall_ws['B23'].value = 'Number of Dropped Calls'
        self.overall_ws['B24'].value = 'Number of Lost Calls'
        self.overall_ws['B25'].value = 'Number of Short Calls'
        self.overall_ws['B26'].value = 'Average Wait Time (Sec)'
        self.overall_ws['B27'].value = 'Average Call Duration (Sec)'
        self.overall_ws['B28'].value = 'Call Abandon Rate (%)'
        self.overall_ws['B29'].value = 'First Call Resolution Rate'
        self.set_border(self.overall_ws, 'B22:C29')

        self.overall_ws['C9'].value  = self.Total_IVR_Session
        self.overall_ws['C10'].value = self.Total_IVR_Session
        self.overall_ws['C11'].value = self.Total_Agent_Session

        # for cell in ['B21', 'B22', 'B23']:
        #     set_font(overall_ws, cell)


    def create_agent_ws(self):
        self.agent_ws.column_dimensions['B'].width = 26
        
        #=========== create row names =====================
        amount_of_agents = len(self.agents)
        row_names_agent = ['Number of Calls Handled','Average Call Duration','Average Wait Time','First Call Resolution Rate','Available Time']
        row_names_overall = ['Overall','Total Number Of Call','Number of Calls Handled','Average Call Duration','Average Wait Time','Number of available agent','Average Available Time']
        row_counter = 3

        for name in row_names_overall:
            self.agent_ws['B'+str(row_counter)] = name
            row_counter+=1
        self.set_border(self.agent_ws,'B3:B'+str(row_counter-1))
        self.set_border(self.agent_ws,'C3:'+get_column_letter(len(self.agent_performance_dict)+2)+str(row_counter-1))

        for agent in range(1,amount_of_agents+1):
            self.agent_ws['B'+str(row_counter)] = f'Agent_{agent}'
            self.set_border(self.agent_ws, 'B'+str(row_counter)+':B'+str(row_counter+len(row_names_agent)))
            self.set_border(self.agent_ws,'C'+str(row_counter)+':'+get_column_letter(len(self.agent_performance_dict)+2)+str(row_counter+len(row_names_agent)))
            #print('B'+str(row_counter)+':B'+str(row_counter+len(row_names_agent)))
            row_counter+=1
            for name in row_names_agent:
                self.agent_ws['B'+str(row_counter)] = name
                row_counter+=1
            
        # ================ set date_time column name =============
        # ================ filling each agent details ==============
        # start at column 3 which is Column C
        column_counter = 2
        for key, agent_ls in self.agent_performance_dict.items():
            #print(f'key: {key}')
            #print(f'agent_ls: {agent_ls}')
            column_counter += 1
            column_letter = get_column_letter(column_counter)
            self.agent_ws.column_dimensions[column_letter].width = 21
            self.agent_ws[column_letter+'2'] = key # set the date_time as column name
            row_counter = 10
            row_name_ls = [
                'Agent',
                'Number_of_Calls_Handled',
                'Average_Call_Duration',
                'Average_Wait_Time',
                'First_Call_Resolution_Rate',
                'Available_Time'
            ]
            for agent in agent_ls:
                for row_name in row_name_ls:
                    self.agent_ws[column_letter+str(row_counter)] = agent[row_name]
                    row_counter +=1
        
        self.draw_plot_at_row = row_counter

        # ================ filling overall details ==============
        column_counter = 2
        for key, item in self.overall_agent_performance_dict.items():
            # print(">>>>>>>>>>>>>>>>>>>>>>")
            # print(f'key: {key}')
            # print(f'item: {item}')
            column_counter += 1
            column_letter = get_column_letter(column_counter)
            
            self.agent_ws[column_letter+'4'] = item['Total_Number_of_Call']
            self.agent_ws[column_letter+'5'] = item['Number_of_Calls_Handled']
            self.agent_ws[column_letter+'6'] = item['Average_Call_Duration']
            self.agent_ws[column_letter+'7'] = item['Average_Wait_Time']
            self.agent_ws[column_letter+'8'] = 'N/A'
            self.agent_ws[column_letter+'9'] = 'N/A'


    def get_24hr_in_every_5min(self):
        # from 00:00:00 to 23:55:00 
        hour = 0
        minute = 0
        my_24 = ['00:00:00']

        while hour <=23:
            if minute >=55:
                hour +=1
                minute=0
            else:
                minute +=5
            hour_ = str(hour)
            minute_ = str(minute)

            if hour <10:
                hour_ = '0'+str(hour)
            if minute<10:
                minute_ = '0'+str(minute)
            my_24.append(hour_+':'+minute_+':00')
        return my_24[:-1]


    def create_utilization_ws(self, filename):
        self.utilization_ws.column_dimensions['B'].width = 20
        for cell in ['C','D','E','F']:
            self.utilization_ws.column_dimensions[cell].width = 15

        # =============== set column names ===============
        self.utilization_ws.merge_cells('C2:F2')
        self.utilization_ws['C2'].alignment = Alignment(horizontal="center", vertical="center")
        self.utilization_ws['C2'].value = 'Utilization'
        self.set_border(self.utilization_ws, 'C2:F2')

        self.utilization_ws.merge_cells('C3:D3')
        self.utilization_ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        self.utilization_ws['C3'].value = 'IVRS'
        self.set_border(self.utilization_ws, 'C3:D3')

        self.utilization_ws.merge_cells('E3:F3')
        self.utilization_ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        self.utilization_ws['E3'].value = 'Agent ACD'
        self.set_border(self.utilization_ws, 'E3:F3')

        my_list = [('Number of call', 'C4'), ('Max throughput', 'D4'),('Number of call', 'E4'),('Max throughput', 'F4')]
        for item in my_list:
            self.utilization_ws[item[1]].alignment = Alignment(horizontal="center", vertical="center")
            self.utilization_ws[item[1]].value = item[0]
            self.utilization_ws[item[1]].border = Border(left=self.thin, right=self.thin, bottom=self.thin, top=self.thin)


        y_ivr_calls = []
        y_agent_calls = []
        x = []

        # ================== set time in rows ==========================
        row_counter = 4
        for datetime, detail in self.datetime_list_5min.items():
            row_counter+=1
            self.utilization_ws['B'+str(row_counter)].value = datetime
            self.utilization_ws['C'+str(row_counter)].value = detail['ivr_calls']
            self.utilization_ws['D'+str(row_counter)].value = detail['ivr_throu']
            self.utilization_ws['E'+str(row_counter)].value = detail['agent_calls']
            self.utilization_ws['F'+str(row_counter)].value = detail['agent_throu']

            x.append(datetime)
            y_ivr_calls.append(detail['ivr_calls'])
            y_agent_calls.append(detail['agent_calls'])

        self.set_border(self.utilization_ws, 'B5:B'+str(row_counter))
        self.set_border(self.utilization_ws, 'C5:D'+str(row_counter))
        self.set_border(self.utilization_ws, 'E5:F'+str(row_counter))


        # =================== plot graph ==============================
        fig = plt.figure()
        plt.plot(x, y_ivr_calls, color='red', marker='o', linestyle='dashed', linewidth=2, markersize=8, label='IVR')
        plt.plot(x, y_agent_calls, color='blue', marker='o', linestyle='dashed', linewidth=2, markersize=8, label='Agent ACD')

        plt.xlabel('Date and Time')
        plt.ylabel('Amount')

        plt.title('Number of call')
        plt.legend()  # to show legend
        #plt.show()

        # rotate the xticks in vertical
        plt.xticks(rotation=90)
        # in order to fully display the long xticks
        plt.tight_layout()

        # set plot height and width
        fig.set_figheight(10)
        #fig.set_figwidth(len(x))
        # fig.set_figheight(10)
        fig.set_figwidth(100)

        plt.savefig('utilization.png')
        print('utilization.png is generated')

        img = drawing.image.Image('utilization.png')
        img.anchor = "I4"
        self.utilization_ws.add_image(img)


    def save_file(self, out_filename):
        self.wb.save(out_filename)


    def date_time_string2object(self, date_time_str):
        # param: date_time_str = "2020-07-15 15:39:08"
        # return date_time_obj = 2020-07-15 15:39:08
        return datetime.strptime(date_time_str, '%Y-%m-%d %H:%M:%S')


    def date_time_object2string(self, date_time_obj):
        # param: date_time_obj = 2020-07-15 15:39:08
        # return date_time_str = "2020-07-15 15:39:08"
        return date_time_obj.strftime("%Y-%m-%d %H:%M:%S")


    def get_date_in_range(self, start_date, end_date):
        '''
        if start_date or end_date have hour minute second, there will be a mistake occurred. (the last one will be missing)
        param - start_date(str): "2008-02-10"
              - end_date  (str): "2008-03-11"
        return
            - datelist(pandas._libs.tslibs.timestamps.Timestamp):  DatetimeIndex(['2008-02-10', '2008-02-11', '2008-02-12', ... ]
                for i in datelist:
                    date = i.strftime("%Y-%m-%d") # to convert in str
        '''

        if len(start_date) > 10:
            start_date = start_date[0:10] # only get the yyyy-mm-dd
        if len(end_date) > 10:
            end_date = end_date[0:10] # only get the yyyy-mm-dd
        # datelist = pd.date_range(start=start_date,end=end_date)  # pandas is quite low for generate the date list

        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d")
        datelist = [start + dt.timedelta(days=x) for x in range(0, (end-start).days+1)]
        return datelist


    def plot_graph(self, out_filename, x, ys, title):
        '''
        ONLY PLOT GRAPH USING MATPLOTLIB
        ys: a list of 3 small lists in which each represents a line, for plot the graph for the overall Total_Number_of_Call, ys will be simple a list of numerics
        x: a list of xticks
        out_filename: the filename to save the plot
        title: the title of the plot
        '''
        fig = plt.figure()
        #print(f'len of ys: {len(ys)}')
        if len(ys) == 0:
            raise Exception("plot_graph - Empty list for ploting graphs")
        elif len(ys) == 3:
            agent_ls = list(self.agents)
            agent_ls.append("OVERALL")
            for i, y in enumerate(ys):
                #print(y)
                # draw a line based on the coordinates
                plt.plot(x, y, color=self.colors[i], marker='o', linestyle='dashed', linewidth=2, markersize=8, label=agent_ls[i])
        else:
            plt.plot(x, ys, color='black', marker='o', linestyle='dashed', linewidth=2, markersize=8, label='OVERALL')

        plt.xlabel('Date and Time')
        #plt.ylabel('Amount')

        plt.title(title)
        plt.legend()  # to show legend
        #plt.show()

        # rotate the xticks in vertical
        plt.xticks(rotation=90)
        # in order to fully display the long xticks
        plt.tight_layout()

        # set plot height and width
        fig.set_figheight(10)
        fig.set_figwidth(len(x))
        # fig.set_figheight(10)
        # fig.set_figwidth(20)

        #plt.savefig("matplotlib_test1.png")
        plt.savefig(f'{out_filename}.png')
        print(f'{out_filename}.png is generated')


    def display_agent_performance_plots(self):
        x = []
        ys_Number_of_Calls_Handled = []
        ys_Average_Call_Duration   = []
        ys_Average_Wait_Time       = []

        for _ in range(len(self.agents)):            
            ys_Number_of_Calls_Handled.append([])
            ys_Average_Call_Duration.append([])
            ys_Average_Wait_Time.append([])

        # print(f'ys_Number_of_Calls_Handled: {ys_Number_of_Calls_Handled}')
        # print(f'ys_Average_Call_Duration: {ys_Average_Call_Duration}')
        # print(f'ys_Average_Wait_Time: {ys_Average_Wait_Time}')

        for date_time, _ in self.agent_performance_dict.items():
            x.append(date_time)

        #============ get data on each agent ===============
        for i, agent_name in enumerate(self.agents):
            for _, agent_ls in self.agent_performance_dict.items():
                for item in agent_ls:
                    if item["Agent"] == agent_name:
                        ys_Number_of_Calls_Handled[i].append(item['Number_of_Calls_Handled'])
                        ys_Average_Call_Duration[i].append(item['Average_Call_Duration'])
                        ys_Average_Wait_Time[i].append(item['Average_Wait_Time'])
                        break

        #============ get data on overall ===============
        y_Total_Number_of_Call = []
        y_overall_Calls_Handled = []
        y_overall_Call_Duration = []
        y_overall_Wait_Time = []
        for _, item in self.overall_agent_performance_dict.items():
            y_Total_Number_of_Call.append(item['Total_Number_of_Call'])  # special case
            y_overall_Calls_Handled.append(item['Number_of_Calls_Handled'])
            y_overall_Call_Duration.append(item['Average_Call_Duration'])
            y_overall_Wait_Time.append(item['Average_Wait_Time'])

        # overall data is placed at the end of each list    
        ys_Number_of_Calls_Handled.append(y_overall_Calls_Handled)
        ys_Average_Call_Duration.append(y_overall_Call_Duration)
        ys_Average_Wait_Time.append(y_overall_Wait_Time)

        #print('display_agent_performance_plots')

        # print(f'len of x: {len(x)}')
        # print(f'len of y_Total_Number_of_Call: {len(y_Total_Number_of_Call)}')
        #print(y_Total_Number_of_Call)
        # print(f'len of ys_Number_of_Calls_Handled[0]: {len(ys_Number_of_Calls_Handled[0])}')
        # print(f'len of ys_Number_of_Calls_Handled[1]: {len(ys_Number_of_Calls_Handled[1])}')
        # print(f'len of ys_Average_Call_Duration[0]: {len(ys_Average_Call_Duration[0])}')
        # print(f'len of ys_Average_Call_Duration[1]: {len(ys_Average_Call_Duration[1])}')
        # for item in zip(x, ys_Average_Call_Duration[0], ys_Average_Call_Duration[1]):
        #     print(item)

        
        self.plot_graph("Number_of_Calls_Handled", x, ys_Number_of_Calls_Handled, "Number_of_Calls_Handled")
        img = drawing.image.Image('Number_of_Calls_Handled.png')
        #img.anchor = "B25"
        img.anchor = "B"+str(self.draw_plot_at_row+3)
        self.agent_ws.add_image(img)

        self.plot_graph("Average_Call_Duration", x, ys_Average_Call_Duration, "Average_Call_Duration")
        img = drawing.image.Image('Average_Call_Duration.png')
        #img.anchor = "B80"
        img.anchor = "B"+str(self.draw_plot_at_row+69)
        self.agent_ws.add_image(img)

        self.plot_graph("Average_Wait_Time", x, ys_Average_Wait_Time, "Average_Wait_Time")
        img = drawing.image.Image('Average_Wait_Time.png')
        #img.anchor = "B140"
        img.anchor = "B"+str(self.draw_plot_at_row+125)
        self.agent_ws.add_image(img)


        ###### for overall amount of calls #####
        self.plot_graph("Total_Number_of_Call", x, y_Total_Number_of_Call, "Total_Number_of_Call")
        img = drawing.image.Image('Total_Number_of_Call.png')
        #img.anchor = "B188"
        img.anchor = "B"+str(self.draw_plot_at_row+181)
        self.agent_ws.add_image(img)




